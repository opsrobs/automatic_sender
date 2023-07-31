import smtplib
import os
import chardet
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

host = "smtp.office365.com"
port = "587"
password = input("Enter your Outlook password: ")

def send_email(email, password, subject, body, attachment):
    email_logged = "robson.flavio20@outlook.com"
    smtp = smtplib.SMTP(host, port)
    smtp.starttls()
    smtp.ehlo()

    print('--------------------------------')
    print(str(email_logged))
    print(str(password))
    smtp.login(email_logged, password)
    msg = MIMEMultipart()
    msg["From"] = email_logged
    msg["To"] = email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    # Attach the PDF file
    with open(attachment, "rb") as f:
        part = MIMEText(f.read(), "application", "pdf")
        part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(attachment)}")
        msg.attach(part)

    # Send the email
    smtp.sendmail(email_logged, email, msg.as_string())

    # Close the SMTP connection
    smtp.quit()

def main():

    xlsx_file_path = input("Enter the path to the XLSX file: ")

    # Get the path to the folder that contains the PDF files
    pdf_folder_path = input("Enter the path to the folder that contains the PDF files: ")
    workbook = openpyxl.load_workbook(xlsx_file_path)
    sheet = workbook.active

    # Find the column index with the header "email"
    email_column_index = None
    for col_idx, cell in enumerate(sheet[1], start=1):  # Assuming the header is in the first row
        if cell.value == "email":
            email_column_index = col_idx
            break

    if email_column_index is None:
        print("Header 'email' not found in the XLSX file.")
        return

    # Get the list of emails from the identified column
    emails = [sheet.cell(row=row_idx, column=email_column_index).value for row_idx in range(2, sheet.max_row + 1)]

    # Loop through the emails
    for email in emails:
        # Get the name of the PDF file
        pdf_file_name = email.split("@")[0] + ".pdf"

        # Get the path to the PDF file
        pdf_file_path = os.path.join(pdf_folder_path, pdf_file_name)

        # Check if the PDF file exists
        if os.path.exists(pdf_file_path):
            # Send the email
            send_email(email, password, "Subject", "Body", pdf_file_path)
        else:
            print(f"Email {email} does not have an attachment and will not be sent.")

if __name__ == "__main__":
    main()
